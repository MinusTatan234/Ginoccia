import time
import serial
import threading
import subprocess
from pathlib import Path
import tkinter as tk
from tkinter import Tk, Canvas, Entry, Button, PhotoImage, ttk, filedialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import numpy as np
from PIL import Image, ImageTk
import cv2
import math
import mediapipe as mp
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
import pandas as pd
import serial.tools.list_ports


_DIR = os.path.dirname(__file__)

# Implementation flags
stop_threads = False
ser = None
activation = False
connect = 0
indicator = None
config_mode_flag = False
study_mode_flag = False
max_angle_right = 0
max_angle_left = 0
position = 0
torque = 0
myoware = 0
set_degr = 0
set_degl = 0
counter = 0
rad = 0
aux_lock = False
dist = 0.22
gravity = 9.81
kg = 0
cnt = 0

data = ""
selected_option = ""
degr = "0"
degl = "0"
text = ""
slider_value = "0"


data_lst = list()
excel_writer = None


def serial_connection():
    # Create serial connection
    global arduino_lock, connect, ser, activation, window, combobox_com, selected_port
    while not stop_threads:
        if ser is None and connect == 1:
            while not stop_threads and connect == 1:
                try:
                    ser = serial.Serial(str(selected_port), 115200, timeout=0.25)
                    if ser is not None:
                        connect = 0
                        indicator.config(bg="green")
                        # window.is_plot_running = True

                except Exception as e:
                    print("The connection was not established ", e)
                    print("Reconnecting...")
                    time.sleep(1)

        elif connect == 2:
            print("Connection close")
            connect = 0

        # time.sleep(0.1)
        if ser is not None:
            activation = True
            arduino_lock = threading.Lock()  # Lock to ensure mutual exclusion between sending and receiving data

            # Starting the reading thread
            reading_thread = threading.Thread(target=read_serial_port, args=(ser,))
            reading_thread.start()

            # Starting the sending thread
            sending_thread = threading.Thread(target=send_data, args=(ser,))
            sending_thread.start()

            # Wait up to all threads end
            reading_thread.join()
            sending_thread.join()

            if activation is True:
                # Close connection
                ser.close()
                print("Serial connection closed")

        elif ser is None:
            print("Serial connection could not be established")
        time.sleep(0.2)


# Function to read serial port
def read_serial_port(arduino):
    global activation, stop_threads, data, position, torque, myoware, counter, button_2,\
        indicator, window, connect, ser, config_mode_flag, study_mode_flag, text, degr, degl, \
        button_4, button_5, button_6, button_7, button_8, button_9, button_10, entry_1, slider, rb_a, rb_b, rbs, \
        combo_box_leg, rad, dist, gravity, kg, combobox_com
    try:
        while not stop_threads and activation is True:
            data = arduino.readline().decode().strip()
            if data:
                counter = 0
                values = data.split(",")
                if len(values) >= 2:
                    rad = abs(float(values[0]))
                    position = (rad * 180) / math.pi
                    torque = abs(float(values[1]))
                    kg = (torque / dist) / gravity
                    myoware = float(values[2])
                print(data)
            else:
                counter = counter + 1
                if counter >= 2:
                    slider.set(0)
                    button_4.config(state="disabled")
                    connect = 2
                    indicator.config(bg="gray")
                    config_mode_flag = False
                    study_mode_flag = False
                    degr = "0"
                    degl = "0"
                    text = ""
                    button_2.config(state="disabled")
                    button_5.config(state="disabled")
                    button_6.config(state="disabled")
                    button_7.config(state="disabled")
                    button_8.config(state="disabled")
                    button_9.config(state="disabled")
                    button_10.config(state="disabled")
                    entry_1.config(state="disabled")
                    slider.config(state="disabled")
                    rb_a.config(state="disabled")
                    rb_b.config(state="disabled")
                    slider.set(0)
                    rbs.set(None)
                    combo_box_leg.current(0)
                    combo_box_leg["state"] = "disabled"
                    combobox_com["state"] = "readonly"
                    combobox_com.current(0)
                    ser.close()
                    ser = None
                    activation = False
                    window.is_plot_running = False

            time.sleep(0.03)  # Short delay to avoid saturating the processor

    except Exception as e:
        print("Error reading the serial port ", e)
        stop_threads = True


# Function to send data by the serial port
def send_data(arduino):
    global arduino_lock, activation, stop_threads, slider_value
    try:
        while not stop_threads and activation is True:
            time.sleep(0.50)

    except Exception as e:
        print("Error sending data ", e)
        stop_threads = True


def interface():
    global indicator, window, connect, ser, activation, config_mode_flag, study_mode_flag, text, degr, degl,\
        button_4, button_5, button_6, button_7, button_8, button_9, button_10, entry_1, slider, rb_a, rb_b, rbs, \
        combo_box_leg, button_2, combobox_com

    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / Path(fr"{str(OUTPUT_PATH)}\assets\frame0")

    def on_window_close():
        global stop_threads
        if activation is True:
            cadena = str(999)
            arduino_lock.acquire()
            time.sleep(0.02)
            ser.write(cadena.encode('ascii'))
            time.sleep(0.02)
            arduino_lock.release()
        stop_threads = True
        window.destroy()
        window.quit()

    def on_enter_pressed(event):
        global text
        text = entry_1.get()
        print("Texto ingresado:", text)
        canvas.itemconfig(tagOrId=cr_patient, text=text)
        entry_1.delete(0,tk.END)

    def show_selected_option(event):
        global selected_option
        selected_option = str(combo_box_leg.get())
        print(selected_option)

    def get_com_ports():
        com_ports = []
        ports = serial.tools.list_ports.comports()
        for port, desc, hwid in sorted(ports):
            com_ports.append(port)
        return com_ports

    def update_combobox(event=None):
        com_ports = get_com_ports()
        combobox_com['values'] = ["None"] + com_ports

    def show_selected_port(event):
        global selected_port
        selected_port = combobox_com.get()
        button_4.config(state="normal")

    def on_slider_changed(event):
        sv = slider.get()

    def send_last_value(event):
        global slider_value
        if selected_option == "Right":
            slider_value = str(int(slider.get()) * -1)
        elif selected_option == "Left":
            slider_value = str(int(slider.get()) * 1)

        if selected_option == "Right" or selected_option == "Left":
            cadena = str(slider_value)
            arduino_lock.acquire()
            time.sleep(0.02)
            ser.write(cadena.encode('ascii'))
            time.sleep(0.02)
            arduino_lock.release()
            print(slider_value)

    def on_radio_btns():
        r_btn_selection = rbs.get()
        if ser is not None:
            if activation is True:
                slider.set(0)
                cadena = str(r_btn_selection)
                arduino_lock.acquire()
                time.sleep(0.02)
                ser.write(cadena.encode('ascii'))
                time.sleep(0.02)
                arduino_lock.release()
                button_9.config(state="normal")
                print(r_btn_selection)

    # Function to set serial connection
    def connect_btn():
        global connect
        if ser is None:
            if activation is False:
                combobox_com["state"] = "disabled"
                button_3.config(state="normal")
                button_4.config(state="disabled")
                button_5.config(state="normal")
                button_10.config(state="normal")
                connect = 1
                slider.set(0)

    def close_btn():
        global connect, ser, activation, config_mode_flag, study_mode_flag, ser, text, degr, degl
        button_4.config(state="normal")
        connect = 2
        indicator.config(bg="gray")
        config_mode_flag = False
        study_mode_flag = False
        degr = "0"
        degl = "0"
        text = ""
        slider.set(0)
        combobox_com["state"] = "readonly"
        combobox_com.current(0)

        if ser is not None:
            if activation is True:
                cadena = str(999)
                arduino_lock.acquire()
                time.sleep(0.02)
                ser.write(cadena.encode('ascii'))
                time.sleep(0.02)
                arduino_lock.release()
                button_2.config(state="disabled")
                button_5.config(state="disabled")
                button_6.config(state="disabled")
                button_7.config(state="disabled")
                button_8.config(state="disabled")
                button_9.config(state="disabled")
                button_10.config(state="disabled")
                entry_1.config(state="disabled")
                slider.config(state="disabled")
                rb_a.config(state="disabled")
                rb_b.config(state="disabled")
                rbs.set(None)
                combo_box_leg.current(0)
                combo_box_leg["state"] = "disabled"

            ser.close()
            ser = None
            activation = False
            window.is_plot_running = False

    def save_btn():
        global degl, degr, max_angle_right, max_angle_left, data_lst, position, set_degr, set_degl
        if config_mode_flag is True and study_mode_flag is False:
            if selected_option == "Right":
                max_angle_right = round(float(set_degr) + position, 2)
                canvas.itemconfig(tagOrId=change_right, text=str(max_angle_right))
            elif selected_option == "Left":
                max_angle_left = round(float(set_degl) + position, 2)
                canvas.itemconfig(tagOrId=change_left, text=str(max_angle_left))
        elif config_mode_flag is False and study_mode_flag is True:
            if selected_option == "Right":
                if max_angle_right >= round(float(set_degr) + position, 2):
                    attr = "No"
                else:
                    attr = "Yes"
                data_lst.append([selected_option, max_angle_right, degr, "weight", attr])

            elif selected_option == "Left":
                if max_angle_left >= round(float(set_degl) + position, 2):
                    attr = "No"
                else:
                    attr = "Yes"
                data_lst.append([selected_option, max_angle_left, degl, "weight", attr])

    def turn_on_motor():
        slider.set(0)
        if ser is not None:
            if activation is True:
                cadena = str(998)
                arduino_lock.acquire()
                time.sleep(0.02)
                ser.write(cadena.encode('ascii'))
                time.sleep(0.02)
                arduino_lock.release()
                button_8.config(state="normal")
                button_9.config(state="disabled")
                rb_a.config(state="disabled")
                rb_b.config(state="disabled")
                slider.config(state="normal")
                combo_box_leg["state"] = "disabled"

    def turn_off_motor():
        slider.set(0)
        if ser is not None:
            if activation is True:
                cadena = str(999)
                arduino_lock.acquire()
                time.sleep(0.02)
                ser.write(cadena.encode('ascii'))
                time.sleep(0.02)
                arduino_lock.release()
                slider.set(0)
                button_8.config(state="disabled")
                rb_a.config(state="normal")
                rb_b.config(state="normal")
                rbs.set(None)
                slider.config(state="disabled")
                combo_box_leg["state"] = "readonly"

    def set_motor_origin():
        global set_degr, set_degl
        if ser is not None:
            if activation is True:
                cadena = str(997)
                arduino_lock.acquire()
                time.sleep(0.02)
                ser.write(cadena.encode('ascii'))
                time.sleep(0.02)
                arduino_lock.release()
                if selected_option == "Right":
                    set_degr = float(degr)
                elif selected_option == "Left":
                    set_degl = float(degl)

    def create_excel():
        global _DIR, text, data_lst, excel_writer
        if data_lst is not []:
            _OUTPUT_DIR = _DIR + "/xls_output/"
            file_name = text + ".xlsx"
            df = pd.DataFrame(data_lst, columns=["Pierna", "Angulo Alcanzado", "Angulo maximo", "Peso aplicado", "Pasa?"])
            excel_writer = pd.ExcelWriter(_OUTPUT_DIR + file_name, engine="openpyxl")
            excel_writer.book = Workbook()
            sheet_names = excel_writer.book.sheetnames
            for sheet_name in sheet_names:
                if sheet_name == "Sheet":
                    del excel_writer.book[sheet_name]
            df.to_excel(excel_writer, sheet_name="Hoja 1", index=False)
            sheet = excel_writer.sheets["Hoja 1"]
            header_fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
            header_font = Font(color="FFFFFF")

            for cell in sheet["1:1"]:
                cell.fill = header_fill
                cell.font = header_font

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value == "Left":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde
                    elif cell.value == "Right":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo

            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 20
            excel_writer.close()
            time.sleep(0.1)
            data_lst = []
            os.system(f'start excel "{_OUTPUT_DIR + file_name}"')
        else:
            print("MALAS NOTICIAS")

    def config_mode_btn():
        global config_mode_flag, study_mode_flag, ser, activation, connect, text, degr, degl, cnt
        # indicator.config(bg="gray")
        if ser is not None:
            if activation is True:

                config_mode_flag = True
                study_mode_flag = False
                degr = "0"
                degl = "0"
                text = ""
                combo_box_leg.current(0)
                slider.set(0)
                canvas.itemconfig(tagOrId=change_right, text=degr)
                canvas.itemconfig(tagOrId=change_left, text=degl)
                canvas.itemconfig(tagOrId=cr_patient, text="")
                # button_3.config(state="disabled")  # close btn disable
                # button_4.config(state="disabled")  # config btn disable
                button_2.config(state="normal")
                button_5.config(state="disabled")  # config btn disable
                button_6.config(state="normal")  # study btn enable
                button_7.config(state="disabled")  # Export to excel btn disable
                button_8.config(state="disabled")
                button_9.config(state="disabled")
                entry_1.config(state="normal")  # Entry text box enable
                slider.config(state="disabled")  # slider disable
                rb_a.config(state="disabled")
                rb_b.config(state="disabled")
                rbs.set(None)
                combo_box_leg["state"] = "readonly"
                combo_box_leg.current(0)
                if cnt != 0:
                    combobox_com.current(0)
                    combobox_com["state"] = "readonly"
                cnt = cnt + 1

                connect = 2
                cadena = str(999)
                arduino_lock.acquire()
                time.sleep(0.02)
                ser.write(cadena.encode('ascii'))
                time.sleep(0.02)
                arduino_lock.release()
            # ser.close()
            # ser = None
            # activation = False
        window.is_plot_running = False

    def study_mode_btn():
        global config_mode_flag, study_mode_flag, text

        if (float(degr) > 0 or float(degl) > 0) and text != "":
            config_mode_flag = False
            study_mode_flag = True
            button_3.config(state="normal")  # close btn enable
            button_5.config(state="normal")  # config btn enable
            button_6.config(state="disabled")  # study btn disable
            button_7.config(state="normal")  # Export to excel btn enable
            entry_1.config(state="disabled")  # Entry text box disable
            # slider.config(state="normal")  # slider enable
            rb_a.config(state="normal")
            rb_b.config(state="normal")
            window.is_plot_running = True

    def open_folder():
        # get the directory within the current script
        path = os.path.dirname(__file__)
        # Open the files explorer in the same project path
        rute = filedialog.askopenfilename(initialdir=path, title="Select archive")
        # If a file is selected
        if rute:
            subprocess.Popen(["start", "", rute], shell=True)

    class PlotUpdater1:
        def __init__(self, root):
            self.root = root
            self.support1 = True
            self.fig, self.ax = plt.subplots(figsize=(6.4, 4))
            self.canvas = FigureCanvasTkAgg(self.fig, master=root)
            self.canvas.get_tk_widget().place(x=14, y=545)  # Ajusta las coordenadas x e y del gráfico
            self.ax.set_xlim(0, 20)  # Modifica el rango del eje x
            self.ax.set_ylim(0, 10)
            self.ax.set_title("EMG")  # Agrega título al gráfico
            self.ax.set_ylabel("Voltage (V)")  # Nombre del eje X
            self.ax.set_xlabel("Time (s)")  # Nombre del eje Y
            self.data = np.random.rand(10)
            self.time = np.arange(10)
            self.line, = self.ax.plot(self.time, self.data)
            self.update_plot()

        def update_plot(self):
            if self.root.is_plot_running:
                self.support1 = True
                self.data[:-1] = self.data[1:]
                self.data[-1] = myoware  # Extraer un solo elemento
                # self.data[-1] = np.random.rand() * 10
                self.time = np.append(self.time[1:], self.time[-1] + 0.2)
                self.line.set_xdata(self.time)
                self.line.set_ydata(self.data)
                self.ax.relim()
                self.ax.autoscale_view()
                # Ajustar los límites del eje x para que vaya de 0 a 20
                self.ax.set_xlim(self.time[0], self.time[-1] + 1)
                self.canvas.draw()
            else:
                if self.support1 is True:
                    self.reset_plot()
                    self.support1 = False

            self.root.after(200, self.update_plot)  # Actualiza cada segundo

        def start_plot(self):
            self.root.is_plot_running = True

        def stop_plot(self):
            self.root.is_plot_running = False

        def reset_plot(self):
            self.ax.cla()  # Limpia el área de trazado
            self.ax.set_xlim(0, 20)  # Reinicia el rango del eje x
            self.ax.set_ylim(0, 10)  # Reinicia el rango del eje y
            self.line, = self.ax.plot([], [])  # Crea una nueva línea vacía
            self.ax.set_title("EMG")  # Agrega título al gráfico
            self.ax.set_ylabel("Voltage (V)")  # Nombre del eje X
            self.ax.set_xlabel("Time (s)")  # Nombre del eje Y
            self.canvas.draw()  # Dibuja la nueva área de trazado
            self.time = np.arange(-9, 1)

    class PlotUpdater2:
        def __init__(self, root):
            self.root = root
            self.support2 = True
            self.fig, self.ax = plt.subplots(figsize=(6.4, 4))
            self.canvas = FigureCanvasTkAgg(self.fig, master=root)
            self.canvas.get_tk_widget().place(x=663, y=545)  # Ajusta las coordenadas x e y del gráfico
            self.ax.set_xlim(0, 20)  # Modifica el rango del eje x
            self.ax.set_ylim(0, 10)
            self.ax.set_title("Applied Wight")  # Agrega título al gráfico
            self.ax.set_ylabel("Kilograms (Kg)")  # Nombre del eje X
            self.ax.set_xlabel("Time (s)")  # Nombre del eje Y
            self.data = np.random.rand(10)
            self.time = np.arange(10)
            self.line, = self.ax.plot(self.time, self.data)
            self.update_plot()

        def update_plot(self):
            if self.root.is_plot_running:
                self.support2 = True
                self.data[:-1] = self.data[1:]
                self.data[-1] = kg  # Extraer un solo elemento
                # self.data[-1] = np.random.rand() * 10
                self.time = np.append(self.time[1:], self.time[-1] + 0.2)
                self.line.set_xdata(self.time)
                self.line.set_ydata(self.data)
                self.ax.relim()
                self.ax.autoscale_view()
                # Ajustar los límites del eje x para que vaya de 0 a 20
                self.ax.set_xlim(self.time[0], self.time[-1] + 1)
                self.canvas.draw()
            else:
                if self.support2 is True:
                    self.reset_plot()
                    self.support2 = False
            self.root.after(200, self.update_plot)  # Actualiza cada segundo

        def start_plot(self):
            self.root.is_plot_running = True

        def stop_plot(self):
            self.root.is_plot_running = False

        def reset_plot(self):
            self.ax.cla()  # Limpia el área de trazado
            self.ax.set_xlim(0, 20)  # Reinicia el rango del eje x
            self.ax.set_ylim(0, 10)  # Reinicia el rango del eje y
            self.line, = self.ax.plot([], [])  # Crea una nueva línea vacía
            self.ax.set_title("Applied Wight")  # Agrega título al gráfico
            self.ax.set_ylabel("Kilograms (Kg)")  # Nombre del eje X
            self.ax.set_xlabel("Time (s)")  # Nombre del eje Y
            self.canvas.draw()  # Dibuja la nueva área de trazado
            self.time = np.arange(-9, 1)

    class CameraViewer:
        def __init__(self, root, position_original=(0, 0), size_original=(640, 480),
                     position_knee=(660, 0), size_knee=(640, 480)):
            self.root = root
            self.position_original = position_original
            self.size_original = size_original
            self.position_knee = position_knee
            self.size_knee = size_knee
            self.video_label_original = tk.Label(root)
            self.video_label_original.place(x=position_original[0], y=position_original[1], width=size_original[0],
                                            height=size_original[1])
            self.video_label_knee = tk.Label(root)
            self.video_label_knee.place(x=position_knee[0], y=position_knee[1], width=size_knee[0], height=size_knee[1])
            self.is_video_playing = False
            self.capture = cv2.VideoCapture(0)  # Abre la cámara
            self.fps = 30  # Tasa de frames por segundo
            self.mp_pose = mp.solutions.pose
            self.pose = self.mp_pose.Pose(static_image_mode=False)

        def right_side(self, results, width, height):
            # X & Y coordinates
            hip_xr = int(results.pose_landmarks.landmark[24].x * width)
            hip_yr = int(results.pose_landmarks.landmark[24].y * height)
            knee_xr = int(results.pose_landmarks.landmark[26].x * width)
            knee_yr = int(results.pose_landmarks.landmark[26].y * height)
            ankle_xr = int(results.pose_landmarks.landmark[28].x * width)
            ankle_yr = int(results.pose_landmarks.landmark[28].y * height)

            # Law of cosines
            Ar = math.sqrt(((ankle_xr - knee_xr) ** 2) + ((ankle_yr - knee_yr) ** 2))
            Br = math.sqrt(((ankle_xr - hip_xr) ** 2) + ((ankle_yr - hip_yr) ** 2))
            Cr = math.sqrt(((hip_xr - knee_xr) ** 2) + ((hip_yr - knee_yr) ** 2))
            try:
                dr = math.degrees(math.acos(((Ar ** 2) - (Br ** 2) + (Cr ** 2)) / (2 * Ar * Cr)))
            except:
                dr = 0
            deg = "{:.1f}".format(dr)

            return deg, hip_xr, hip_yr, knee_xr, knee_yr, ankle_xr, ankle_yr

        def left_side(self, results, width, height):
            # X & Y coordinates

            hip_xl = int(results.pose_landmarks.landmark[23].x * width)
            hip_yl = int(results.pose_landmarks.landmark[23].y * height)
            knee_xl = int(results.pose_landmarks.landmark[25].x * width)
            knee_yl = int(results.pose_landmarks.landmark[25].y * height)
            ankle_xl = int(results.pose_landmarks.landmark[27].x * width)
            ankle_yl = int(results.pose_landmarks.landmark[27].y * height)

            # Law of cosines
            Al = math.sqrt(((ankle_xl - knee_xl) ** 2) + ((ankle_yl - knee_yl) ** 2))
            Bl = math.sqrt(((ankle_xl - hip_xl) ** 2) + ((ankle_yl - hip_yl) ** 2))
            Cl = math.sqrt(((hip_xl - knee_xl) ** 2) + ((hip_yl - knee_yl) ** 2))
            try:
                dl = math.degrees(math.acos(((Al ** 2) - (Bl ** 2) + (Cl ** 2)) / (2 * Al * Cl)))
            except:
                dl = 0
            deg = "{:.1f}".format(dl)

            return deg, hip_xl, hip_yl, knee_xl, knee_yl, ankle_xl, ankle_yl

        def visualization(self, deg, hip_x, hip_y, knee_x, knee_y, ankle_x, ankle_y, frame):
            aux_img = np.zeros(frame.shape, np.uint8)
            cv2.line(aux_img, (hip_x, hip_y), (knee_x, knee_y), (255, 255, 255), 2)
            cv2.line(aux_img, (knee_x, knee_y), (ankle_x, ankle_y), (255, 255, 255), 2)
            cv2.line(aux_img, (hip_x, hip_y), (ankle_x, ankle_y), (255, 255, 255), 2)

            contours_r = np.array([[hip_x, hip_y], [knee_x, knee_y], [ankle_x, ankle_y]])

            cv2.fillPoly(aux_img, pts=[contours_r], color=(255, 128, 0))
            # Output frame with the effects applied
            output = cv2.addWeighted(frame, 1, aux_img, 0.8, 0)

            cv2.circle(output, (hip_x, hip_y), 6, (0, 255, 255), -1)
            cv2.circle(output, (knee_x, knee_y), 6, (0, 255, 255), -1)
            cv2.circle(output, (ankle_x, ankle_y), 6, (0, 255, 255), -1)

            cv2.putText(output, str(deg), (knee_x + 30, knee_y), 1, 1.5, (128, 0, 250), 2)
            frame2_rgb = cv2.cvtColor(output, cv2.COLOR_BGR2RGB)
            return frame2_rgb

        def detect_knee(self, frame):
            global degl, degr, set_degr, set_degl, position
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            height, width, _ = frame_rgb.shape
            results = self.pose.process(frame_rgb)
            if results.pose_landmarks:
                if selected_option == "Right":
                    degr, hip_xr, hip_yr, knee_xr, knee_yr, ankle_xr, ankle_yr = self.right_side(results, width, height)
                    # final_degr = float(set_degr) + ((float(position) * 180) / math.pi)
                    frame2_rgb = self.visualization(round(float(set_degr) + position, 2), hip_xr, hip_yr, knee_xr, knee_yr, ankle_xr, ankle_yr, frame)
                    return frame2_rgb
                elif selected_option == "Left":
                    # X & Y coordinates
                    degl, hip_xl, hip_yl, knee_xl, knee_yl, ankle_xl, ankle_yl = self.left_side(results, width, height)
                    # Visualization
                    # final_degl = float(set_degl) + ((float(position) * 180)/math.pi)
                    frame2_rgb = self.visualization(round(float(set_degl) + position, 2), hip_xl, hip_yl, knee_xl, knee_yl, ankle_xl, ankle_yl, frame)
                    return frame2_rgb
            return frame_rgb

        def update_video(self):
            ret, frame = self.capture.read()  # Lee un fotograma de la cámara
            if ret and self.root.is_video_playing:
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img_original = Image.fromarray(frame_rgb)
                img_original = img_original.resize(self.size_original)  # Redimensiona la imagen al tamaño deseado
                imgtk_original = ImageTk.PhotoImage(image=img_original)
                self.video_label_original.imgtk = imgtk_original
                self.video_label_original.configure(image=imgtk_original)
                frame_with_knee = self.detect_knee(frame)
                frame_with_knee_rgb = frame_with_knee
                img_knee = Image.fromarray(frame_with_knee_rgb)
                img_knee = img_knee.resize(self.size_knee)  # Redimensiona la imagen al tamaño deseado
                imgtk_knee = ImageTk.PhotoImage(image=img_knee)
                self.video_label_knee.imgtk = imgtk_knee
                self.video_label_knee.configure(image=imgtk_knee)
            self.root.after(int(1000 / self.fps), self.update_video)  # Actualiza con la tasa de frames por segundo

        def start_video(self):
            self.is_video_playing = True

        def stop_video(self):
            self.is_video_playing = False

    def relative_to_assets(path: str) -> Path:
        return ASSETS_PATH / Path(path)

    global window
    window = Tk()
    window.geometry("1900x950")
    window.configure(bg="white")
    window.protocol("WM_DELETE_WINDOW", on_window_close)

    canvas = Canvas(
        window,
        bg="#FFFFFF",
        height=915,
        width=1574,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas.place(x=0, y=0)
    image_image_1 = PhotoImage(
        file=relative_to_assets("image_1.png"))
    image_1 = canvas.create_image(
        334.0,
        351.0,
        image=image_image_1
    )

    image_image_2 = PhotoImage(
        file=relative_to_assets("image_2.png"))
    image_2 = canvas.create_image(
        983.0,
        728.0,
        image=image_image_2
    )

    image_image_3 = PhotoImage(
        file=relative_to_assets("image_3.png"))
    image_3 = canvas.create_image(
        983.0,
        351.0,
        image=image_image_3
    )

    image_image_4 = PhotoImage(
        file=relative_to_assets("image_4.png"))
    image_4 = canvas.create_image(
        334.0,
        728.0,
        image=image_image_4
    )

    button_image_1 = PhotoImage(
        file=relative_to_assets("button_1.png"))
    button_1 = Button(
        image=button_image_1,
        borderwidth=0,
        highlightthickness=0,
        command=open_folder,
        relief="flat"
    )
    button_1.place(
        x=1329.0,
        y=848.0,
        width=230.0,
        height=60.0
    )

    button_image_2 = PhotoImage(
        file=relative_to_assets("button_2.png"))
    button_2 = Button(
        image=button_image_2,
        borderwidth=0,
        highlightthickness=0,
        command=save_btn,
        relief="flat",
        state="disabled"
    )
    button_2.place(
        x=1329.0,
        y=698.0,
        width=230.0,
        height=60.0
    )

    button_image_3 = PhotoImage(
        file=relative_to_assets("button_3.png"))
    button_3 = Button(
        image=button_image_3,
        borderwidth=0,
        highlightthickness=0,
        command=close_btn,
        relief="flat",
        # state="disabled"
    )
    button_3.place(
        x=1030.0,
        y=90.0,
        width=230.0,
        height=60.0
    )

    button_image_4 = PhotoImage(
        file=relative_to_assets("button_4.png"))
    button_4 = Button(
        image=button_image_4,
        borderwidth=0,
        highlightthickness=0,
        command=connect_btn,
        relief="flat",
        state="disabled"
        # state="disabled"
    )
    button_4.place(
        x=1030.0,
        y=9.0,
        width=230.0,
        height=60.0
    )

    button_image_5 = PhotoImage(
        file=relative_to_assets("button_5.png"))
    button_5 = Button(
        window,
        image=button_image_5,
        borderwidth=0,
        highlightthickness=0,
        command=config_mode_btn,
        relief="flat",
        state="disabled"
    )
    button_5.place(
        x=23.0,
        y=9.0,
        width=230.0,
        height=60.0
    )

    button_image_6 = PhotoImage(
        file=relative_to_assets("button_6.png"))
    button_6 = Button(
        image=button_image_6,
        borderwidth=0,
        highlightthickness=0,
        command=study_mode_btn,
        relief="flat",
        state="disabled"
    )
    button_6.place(
        x=23.0,
        y=90.0,
        width=230.0,
        height=60.0
    )
    button_image_7 = PhotoImage(
        file=relative_to_assets("button_7.png"))
    button_7 = Button(
        image=button_image_7,
        borderwidth=0,
        highlightthickness=0,
        command=create_excel,
        relief="flat",
        state="disabled"
    )
    button_7.place(
        x=1330.0,
        y=773.0,
        width=230.0,
        height=60.0
    )
    button_image_8 = PhotoImage(
        file=relative_to_assets("button_8.png"))
    button_8 = Button(
        image=button_image_8,
        borderwidth=0,
        highlightthickness=0,
        command=turn_off_motor,
        relief="flat",
        state="disabled"
    )
    button_8.place(
        x=1585.0,
        y=848.0,
        width=230.0,
        height=60.0
    )
    button_image_9 = PhotoImage(
        file=relative_to_assets("button_9.png"))
    button_9 = Button(
        image=button_image_9,
        borderwidth=0,
        highlightthickness=0,
        command=turn_on_motor,
        relief="flat",
        state="disabled"
    )
    button_9.place(
        x=1585.0,
        y=699.0,
        width=230.0,
        height=60.0
    )
    button_image_10 = PhotoImage(
        file=relative_to_assets("button_10.png"))
    button_10 = Button(
        image=button_image_10,
        borderwidth=0,
        highlightthickness=0,
        command=set_motor_origin,
        relief="flat",
        state="disabled"
    )
    button_10.place(
        x=1585.0,
        y=774.0,
        width=230.0,
        height=60.0
    )
    refresh_button = ttk.Button(window, text="Update", command=update_combobox)
    refresh_button.place(x=780, y=120)
    entry_image_1 = PhotoImage(
        file=relative_to_assets("entry_1.png"))
    entry_bg_1 = canvas.create_image(
        1445.0,
        113.0,
        image=entry_image_1
    )
    entry_1 = Entry(
        bd=0,
        bg="#D9D9D9",
        fg="#000716",
        highlightthickness=0,
        state="disabled"
    )
    entry_1.bind('<Return>', on_enter_pressed)
    entry_1.place(
        x=1350.0,
        y=85.0,
        width=190.0,
        height=58.0
    )

    txt_image_1 = PhotoImage(
        file=relative_to_assets("entry_1.png"))
    entry_bg_1 = canvas.create_image(
        1445.0,
        224.0,
        image=entry_image_1
    )
    canvas.create_text(
        1338.0,
        60,
        anchor="nw",
        text="Name of patient",
        fill="#000000",
        font=("Inter Black", 13 * -1)
    )
    canvas.create_text(
        1338.0,
        171.0,
        anchor="nw",
        text="Current patient",
        fill="#000000",
        font=("Inter Black", 13 * -1)
    )
    canvas.create_text(
        275.0,
        12.0,
        anchor="nw",
        text="Maximum configured angle of the left leg",
        fill="#000000",
        font=("Inter Black", 13 * -1)
    )

    canvas.create_text(
        525.0,
        14.0,
        anchor="nw",
        text="Maximum configured angle of the right leg",
        fill="#000000",
        font=("Inter Black", 13 * -1)
    )
    canvas.create_text(
        780.0,
        14.0,
        anchor="nw",
        text="Port to set connection",
        fill="#000000",
        font=("Inter Black", 13 * -1)
    )
    image_image_5 = PhotoImage(
        file=relative_to_assets("image_5.png"))
    image_5 = canvas.create_image(
        375.0,
        83.0,
        image=image_image_5
    )

    image_image_6 = PhotoImage(
        file=relative_to_assets("image_6.png"))
    image_6 = canvas.create_image(
        625.0,
        83.0,
        image=image_image_6
    )

    change_right = canvas.create_text( #right
        615.0,
        70.0,
        anchor="nw",
        text="0",
        fill="#000000",
        font=("Inter Black", 24 * -1)
    )

    change_left = canvas.create_text( #left
        360.0,
        70.0,
        anchor="nw",
        text="0",
        fill="#000000",
        font=("Inter Black", 24 * -1)
    )
    cr_patient = canvas.create_text(
        1350.0,
        212.0,
        anchor="nw",
        text="",
        fill="#000000",
        font=("Inter Black", 24 * -1)
    )

    canvas.create_text(
        1338.0,
        300.0,
        anchor="nw",
        text="Choose Leg Side",
        fill="#000000",
        font=("Inter Black", 13 * -1)
    )

    canvas.create_text(
        1338,
        570,
        anchor="nw",
        text="Strength",
        fill="#000000",
        font=("Inter Black", 13 * -1)
    )
    canvas.create_text(
        1338,
        440,
        anchor="nw",
        text="Working Mode",
        fill="#000000",
        font=("Inter Black", 13 * -1)
    )
    indicator = tk.Label(
        window,
        width=5,
        height=2,
        bg="gray"
    )
    indicator.place(x=1265, y=60)
    # Add ComboBox
    options = ["None", "Right", "Left"]
    combo_box_leg = ttk.Combobox(
        window,
        values=options,
        state="readonly",
    )
    combo_box_leg.place(
        x=1338,
        y=320,
        width=220.0,
        height=60.0
    )
    combo_box_leg["state"] = "disabled"
    combo_box_leg.current(0)
    # Bind function show_selected_option to event <<ComboboxSelected>>
    combo_box_leg.bind("<<ComboboxSelected>>", show_selected_option)

    combo_box_leg.configure(
        font=(
            "Inter",
            24,
            "bold"
        ),
        justify="center",
    )

    # Crear el combobox
    options = get_com_ports()
    options = ["None"] + options
    combobox_com = ttk.Combobox(
        window,
        values=options,
        state="readonly",
    )
    combobox_com.place(x=780, y=53, width=220.0, height=60.0)
    combobox_com["state"] = "readonly"
    combobox_com.current(0)
    combobox_com.bind("<<ComboboxSelected>>", show_selected_port)

    combobox_com.configure(
        font=("Inter", 24, "bold"),
        justify="center",
    )

    # Add a slider from 0 to 12
    slider = tk.Scale(
        window,
        from_=0,
        to=12,
        tickinterval=1,
        orient=tk.HORIZONTAL,
        command=on_slider_changed,
        state="disabled"
    )

    slider.place(
        x=1338.0,
        y=600.0,
        width=260,
        height=60
    )
    # lbl = tk.Label(window, text="Selected value: 0")
    # lbl.pack()
    slider.bind("<ButtonRelease-1>", send_last_value)

    rbs = tk.StringVar()
    rbs.set(None)

    rb_a = tk.Radiobutton(window, text="Spring Mode", variable=rbs, value="a", command=on_radio_btns,
                          state="disabled", font=("Arial", 18, "bold"), anchor='w')
    rb_b = tk.Radiobutton(window, text="Constant Torque", variable=rbs, value="b", command=on_radio_btns,
                          state="disabled", font=("Arial", 18, "bold"), anchor='w')
    rb_a.place(x=1338, y=460, width=260, height=30)
    rb_b.place(x=1338, y=500, width=260, height=30)

    window.is_plot_running = False
    window.is_video_playing = True
    size_original = (640, 360)
    size_knee = (640, 360)
    plot_updater1 = PlotUpdater1(window)
    plot_updater2 = PlotUpdater2(window)
    camera_viewer = CameraViewer(window, position_original=(14, 171), size_original=size_original,
                                 position_knee=(661, 171), size_knee=size_knee)

    camera_viewer.start_video()
    camera_viewer.update_video()
    window.resizable(False, False)
    window.mainloop()


# Starting the interface

start_connection = threading.Thread(target=serial_connection)
start_connection.start()

interface()

start_connection.join()
