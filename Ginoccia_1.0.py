import cv2
import mediapipe as mp
import numpy as np
import math
import threading
import time
import serial
import os
import pandas as pd
import tkinter as tk
from pynput import keyboard as kb
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

global _DIR
_DIR = os.path.dirname(__file__)
print(_DIR)
arduino = None
ctu = 0
# list_of_weights = [0, 0, 0, 137.5, 275, 480, 600, 650, 690, 825, 915, 1015, 1190]
list_of_weights = [0, 0, 0, 200, 300, 500, 600, 650, 750, 850, 900, 1100, 1500]

if not os.path.exists(_DIR + "/xls_output"):
    os.makedirs(_DIR + "/xls_output")


def pulsa(tecla):
    global s
    global itr
    global down
    global up
    global ctu
    global arduino
    global lock
    global lst
    global mem
    global max_angle_left
    global max_angle_right
    global weight
    global degl
    global degr
    global angles
    global mar
    global backup
    if tecla == kb.KeyCode.from_char('c') and s != 0:
        print('config mode')
        backup = 0
        angles = []
        s = 0
        itr = 0
        down = 0
        up = 0
        ctu = 0
    elif tecla == kb.KeyCode.from_char('l') and s != 0:
        print('left')
        backup = 0
        angles = []
        for i in range(24):
            time.sleep(0.02)
            cadena = '-'
            lock = 1
            time.sleep(0.02)
            arduino.write(cadena.encode('ascii'))
            time.sleep(0.02)
            lock = 0
        time.sleep(0.02)
        s = 1
        itr = 0
        down = 0
        up = 0
        ctu = 0
        mem = 0
    elif tecla == kb.KeyCode.from_char('r') and s != 0:
        print('right')
        backup = 0
        angles = []
        for i in range(24):
            time.sleep(0.02)
            cadena = '-'
            lock = 1
            time.sleep(0.02)
            arduino.write(cadena.encode('ascii'))
            time.sleep(0.02)
            lock = 0
        time.sleep(0.02)
        s = 2
        itr = 0
        down = 0
        up = 0
        ctu = 0
        mem = 0
    elif tecla == kb.KeyCode.from_char('+') and s != 0:
        cadena = '+'
        lock = 1
        time.sleep(0.02)
        arduino.write(cadena.encode('ascii'))
        time.sleep(0.02)
        lock = 0
        backup = 0
    elif tecla == kb.KeyCode.from_char('-') and s != 0:
        cadena = '-'
        lock = 1
        time.sleep(0.02)
        arduino.write(cadena.encode('ascii'))
        time.sleep(0.02)
        lock = 0
        backup = 0
    elif tecla == kb.KeyCode.from_char('s') and s != 0:
        if mem != ctu:
            mem = ctu
            if s == 1:
                lst.append(['left', backup, max_angle_left, weight, 'Si'])
            elif s == 2:
                lst.append(['right', backup, max_angle_right, weight, 'Si'])
        else:
            if s == 1:
                lst.append(['left', backup, max_angle_left, weight, 'No'])
            elif s == 2:
                lst.append(['right', backup, max_angle_right, weight, 'No'])
        print(lst)
        backup = 0


def visualization(deg, hip_x, hip_y, knee_x, knee_y, ankle_x, ankle_y, frame):
    global up
    global down
    global max_angle_left
    global max_angle_right
    global s
    global ctu
    global angles
    global vlock
    global mar
    global backup
    aux_img = np.zeros(frame.shape, np.uint8)
    cv2.line(aux_img, (hip_x, hip_y), (knee_x, knee_y), (255, 255, 255), 2)
    cv2.line(aux_img, (knee_x, knee_y), (ankle_x, ankle_y), (255, 255, 255), 2)
    cv2.line(aux_img, (hip_x, hip_y), (ankle_x, ankle_y), (255, 255, 255), 2)

    contours_r = np.array([[hip_x, hip_y], [knee_x, knee_y], [ankle_x, ankle_y]])

    cv2.fillPoly(aux_img, pts=[contours_r], color=(255, 128, 0))
    # Output frame with the effects applied
    output = cv2.addWeighted(frame, 1, aux_img, 0.8, 0)

    cv2.circle(output, (hip_x, hip_y), 6, (0, 255, 255), -1)
    cv2.circle(output, (knee_x, knee_y), 6, (0, 255, 255), -1)
    cv2.circle(output, (ankle_x, ankle_y), 6, (0, 255, 255), -1)

    cv2.putText(output, str(deg), (knee_x + 30, knee_y), 1, 1.5, (128, 0, 250), 2)
    cv2.putText(output, str(ctu), (50, 50), 1, 3.0, (0, 0, 0), 2)
    frame2_rgb = cv2.cvtColor(output, cv2.COLOR_BGR2RGB)
    if float(deg) >= 110 and (s == 1 or s == 2):
        vlock = 0
        angles.append(deg)

    elif float(deg) < 110 and (s == 1 or s == 2) and vlock == 0:
        print(angles)
        try:
            mar = float(max(angles))
            if backup < mar:
                backup = mar
        except:
            pass
        angles = []
        vlock = 1

    if s == 1:
        if (float(deg) >= 85) and (float(deg) <= 95) and down == 0 and up == 0:
            down = 1
            print("down ", down)
        elif (float(deg) >= float(max_angle_left) - 5) and (float(deg) <= float(max_angle_left) + 5) and down == 1 and up == 0:
            up = 1
            print("up ", up)
        elif (float(deg) >= 85) and (float(deg) <= 95) and down == 1 and up == 1:
            down = 0
            up = 0
            ctu = ctu + 1
            print("ctu ", ctu)
    elif s == 2:
        if (float(deg) >= 85) and (float(deg) <= 95) and down == 0 and up == 0:
            down = 1
            print("down ", down)
        elif (float(deg) >= float(max_angle_right) - 5) and (float(deg) <= float(max_angle_right) + 5) and down == 1 and up == 0:
            up = 1
            print("up ", up)
        elif (float(deg) >= 85) and (float(deg) <= 95) and down == 1 and up == 1:
            down = 0
            up = 0
            ctu = ctu + 1
            print("ctu ", ctu)

    return frame2_rgb


def right_side(results, width, height):
    # X & Y coordinates
    hip_xr = int(results.pose_landmarks.landmark[24].x * width)
    hip_yr = int(results.pose_landmarks.landmark[24].y * height)
    knee_xr = int(results.pose_landmarks.landmark[26].x * width)
    knee_yr = int(results.pose_landmarks.landmark[26].y * height)
    ankle_xr = int(results.pose_landmarks.landmark[28].x * width)
    ankle_yr = int(results.pose_landmarks.landmark[28].y * height)

    # Law of cosines
    Ar = math.sqrt(((ankle_xr - knee_xr) ** 2) + ((ankle_yr - knee_yr) ** 2))
    Br = math.sqrt(((ankle_xr - hip_xr) ** 2) + ((ankle_yr - hip_yr) ** 2))
    Cr = math.sqrt(((hip_xr - knee_xr) ** 2) + ((hip_yr - knee_yr) ** 2))
    try:
        dr = math.degrees(math.acos(((Ar ** 2) - (Br ** 2) + (Cr ** 2)) / (2 * Ar * Cr)))
    except:
        dr = 0
    deg = "{:.1f}".format(dr)

    return deg, hip_xr, hip_yr, knee_xr, knee_yr, ankle_xr, ankle_yr


def left_side(results, width, height):
    # X & Y coordinates
    hip_xl = int(results.pose_landmarks.landmark[23].x * width)
    hip_yl = int(results.pose_landmarks.landmark[23].y * height)
    knee_xl = int(results.pose_landmarks.landmark[25].x * width)
    knee_yl = int(results.pose_landmarks.landmark[25].y * height)
    ankle_xl = int(results.pose_landmarks.landmark[27].x * width)
    ankle_yl = int(results.pose_landmarks.landmark[27].y * height)

    # Law of cosines
    Al = math.sqrt(((ankle_xl - knee_xl) ** 2) + ((ankle_yl - knee_yl) ** 2))
    Bl = math.sqrt(((ankle_xl - hip_xl) ** 2) + ((ankle_yl - hip_yl) ** 2))
    Cl = math.sqrt(((hip_xl - knee_xl) ** 2) + ((hip_yl - knee_yl) ** 2))
    try:
        dl = math.degrees(math.acos(((Al ** 2) - (Bl ** 2) + (Cl ** 2)) / (2 * Al * Cl)))
    except:
        dl = 0
    deg = "{:.1f}".format(dl)

    return deg, hip_xl, hip_yl, knee_xl, knee_yl, ankle_xl, ankle_yl


def create_excel(nme, lst=None):
    global _DIR
    if lst is not None:
        _OUTPUT_DIR = _DIR + "/xls_output/"
        file_name = nme+".xlsx"
        df = pd.DataFrame(lst, columns=["Pierna", "Angulo Alcanzado", "Angulo maximo", "Peso aplicado", "Pasa?"])
        excel_writer = pd.ExcelWriter(_OUTPUT_DIR + file_name, engine="openpyxl")
        excel_writer.book = Workbook()
        sheet_names = excel_writer.book.sheetnames
        for sheet_name in sheet_names:
            if sheet_name == "Sheet":
                del excel_writer.book[sheet_name]
        df.to_excel(excel_writer, sheet_name="Hoja 1", index=False)
        sheet = excel_writer.sheets["Hoja 1"]
        header_fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        header_font = Font(color="FFFFFF")

        for cell in sheet["1:1"]:
            cell.fill = header_fill
            cell.font = header_font

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == "left":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Verde
                elif cell.value == "right":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rojo

        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        excel_writer.save()
        os.system(f'start excel "{_OUTPUT_DIR + file_name}"')
    else:
        print("MALAS NOTICIAS")
        return None


class TimerThread(threading.Thread):
    def __init__(self):
        super().__init__()
        self.start_time = None
        self.end_time = None
        self.exit_event = threading.Event()
        self.time_elapsed = 0

    def run(self):
        self.start_time = time.time()
        while not self.exit_event.is_set():
            self.time_elapsed = time.time() - self.start_time
            time.sleep(1)  # update time every second

    def stop(self):
        self.exit_event.set()
        self.end_time = time.time()

    def get_elapsed_time(self):
        if self.end_time is not None:
            return self.end_time - self.start_time
        else:
            return self.time_elapsed


# THREADS
def thread1():
    global s
    global itr
    global down
    global up
    global max_angle_left
    global max_angle_right
    global lock
    global arduino
    global mem
    global lst
    global weight
    global degr
    global degl
    global ctu
    global angles
    global vlock
    print("Program is running for the first time, config mode is launched")
    s = 0
    itr = 0
    left = 0
    right = 0
    down = 0
    up = 0
    lock = 0
    available = 0
    mem = 0
    weight = 0
    l = 0
    lst = []
    angles = []
    counter = 0
    vlock = 0
    max_angle_left = 0
    max_angle_right = 0
    pc = 0
    # variable that stores an API tool to draw lines over images
    # mp_drawing = mp.solutions.drawing_utils
    # variable that stores an API tool to detect a human body
    mp_pose = mp.solutions.pose

    frame2_rgb = np.zeros((360, 640, 3), np.uint8)

    root = tk.Tk()
    root.title("Frames de OpenCV en Tkinter")

    frame_video = tk.Frame(root)
    frame_video.pack()
    # Crear las etiquetas para mostrar los frames
    label1 = tk.Label(frame_video)
    label1.pack(side="left")

    label2 = tk.Label(frame_video)
    label2.pack(side="right")

    fig1 = Figure(figsize=(5, 4), dpi=100)
    ax1 = fig1.add_subplot(111)
    line1, = ax1.plot([], [])
    ax1.set_title("Señal miolectrica")
    ax1.set_xlabel("Tiempo (s)")
    ax1.set_ylabel("Voltaje")
    ax1.set_xlim(0, 300)
    ax1.set_ylim(0, 6)

    canvas1 = FigureCanvasTkAgg(fig1, master=root)  # Convierte la figura en un widget de Tkinter
    canvas1.draw()
    canvas_widget1 = canvas1.get_tk_widget()
    canvas_widget1.pack(side="left", fill="both", expand=True)

    fig2 = Figure(figsize=(5, 4), dpi=100)
    ax2 = fig2.add_subplot(111)
    line2, = ax2.plot([], [])
    ax2.set_title("Strength lvl (gr)")
    ax2.set_xlabel("Tiempo (s)")
    ax2.set_ylabel("Strength")
    ax2.set_xlim(0, 300)
    ax2.set_ylim(0, 1600)

    canvas2 = FigureCanvasTkAgg(fig2, master=root)  # Convierte la figura en un widget de Tkinter
    canvas2.draw()
    canvas_widget2 = canvas2.get_tk_widget()
    canvas_widget2.pack(side="right", fill="both", expand=True)

    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    # cv2.namedWindow('aux', cv2.WINDOW_NORMAL)
    # cv2.namedWindow('Frame', cv2.WINDOW_NORMAL)

    with mp_pose.Pose(static_image_mode=False) as pose:  # Tool config
        while True:
            if available == 1 and s != 0:
                if arduino.in_waiting > 0:
                    if lock == 0:
                        data = str(arduino.readline().decode().strip())
                        try:
                            d = data.split(",")
                            select = int(float(d[1]))
                            weight = list_of_weights[select]
                            voltage = float(d[2])
                            line1.set_xdata(np.append(line1.get_xdata(), counter))
                            line1.set_ydata(np.append(line1.get_ydata(), voltage))
                            ax1.relim()
                            ax1.set_xlim(max(0, counter - 10), max(10, counter + 10))
                            fig1.canvas.draw_idle()
                        except:
                            continue

                        line2.set_xdata(np.append(line2.get_xdata(), counter))
                        line2.set_ydata(np.append(line2.get_ydata(), weight))
                        ax2.relim()
                        ax2.set_xlim(max(0, counter - 10), max(10, counter + 10))
                        fig2.canvas.draw_idle()

                        # print(f' Strength {list_of_weights[select]} gr', f' Voltage {d[2]} V')
                        counter += 0.2
            ret, frame = cap.read()
            if not ret:
                break
            height, width, _ = frame.shape  # From "frame" extract height and width
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)  # Transform frame format from BGR to RGB
            img = Image.fromarray(frame_rgb)
            img = ImageTk.PhotoImage(image=img)

            label1.config(image=img)
            label1.img = img
            results = pose.process(frame_rgb)  # Once a frame is captured, the mediapipe method and proces is applied

            if results.pose_landmarks is not None:  # Checks if truly exists "pose landmarks"
                if s == 2:  # Right
                    # X & Y coordinates
                    degr, hip_xr, hip_yr, knee_xr, knee_yr, ankle_xr, ankle_yr = right_side(results, width, height)
                    # Visualization
                    frame2_rgb = visualization(degr, hip_xr, hip_yr, knee_xr, knee_yr, ankle_xr, ankle_yr, frame)
                elif s == 1:  # Left
                    # X & Y coordinates
                    degl, hip_xl, hip_yl, knee_xl, knee_yl, ankle_xl, ankle_yl = left_side(results, width, height)
                    # Visualization
                    frame2_rgb = visualization(degl, hip_xl, hip_yl, knee_xl, knee_yl, ankle_xl, ankle_yl, frame)

                elif s == 0:  # configuration
                    if available == 0 and l == 0:
                        nme = str(input("Patient's Name: "))
                        l = 1
                    elif available == 1:
                        for i in range(24):
                            time.sleep(0.02)
                            cadena = '-'
                            lock = 1
                            time.sleep(0.02)
                            arduino.write(cadena.encode('ascii'))
                            time.sleep(0.02)
                            lock = 0
                        time.sleep(0.02)
                        arduino.close()
                        create_excel(nme, lst)
                        lst = []
                        available = 0
                        nme = str(input("Patient's Name: "))
                        l = 1

                    if left == 0 and right == 0 and itr == 0:
                        while True:
                            side = str(input("Choose the side of the leg you want to set: ")).lower()
                            if side == 'l' or side == 'r':
                                break
                            else:
                                print('Not valid option, try again')
                                time.sleep(1)
                        elapsed_time = None
                    elif left == 1 and right == 0 and itr == 0:
                        while True:
                            side = str(input("Left leg has been calibrated, press r to configure right leg, l to "
                                             "reconfigure left leg or c to straight forward: ")).lower()
                            if side == 'l' or side == 'r':
                                break
                            elif side == 'c':
                                left = 1
                                right = 1
                                pc = 1
                                break
                            else:
                                print('Not valid option, try again')
                                time.sleep(1)
                        elapsed_time = None
                    elif left == 0 and right == 1 and itr == 0:
                        while True:
                            side = str(input("Right leg has been calibrated, press l to configure left leg, r to "
                                             "reconfigure right leg or c to straight forward: ")).lower()
                            if side == 'l' or side == 'r':
                                break
                            elif side == 'c':
                                left = 1
                                right = 1
                                pc = 2
                                break
                            else:
                                print('Not valid option, try again')
                                time.sleep(1)
                        elapsed_time = None
                    elif left == 1 and right == 1 and itr == 0:
                        while True:
                            if pc == 0:
                                side = str(input("Left and right legs have been calibrated, press c to leave from config "
                                                 "mode or press whether l or r to reconfigure any leg: ")).lower()
                            elif pc == 1:
                                side = str(input("Left lef has been calibrated, press c to leave from config ")).lower()
                            elif pc == 2:
                                side = str(input("Right leg has been calibrated, press c to leave from config ")).lower()
                            pc = 0
                            if side == 'l' or side == 'r' or side == 'c':
                                break
                            else:
                                print('Not valid option, try again')
                                time.sleep(1)
                        if side == 'c':
                            print("left = ", max_angle_left)
                            print("right = ", max_angle_right)
                            print("press l or r keys to start the analysis, or c to restart configurations")
                            s = -1
                            left = 0
                            right = 0
                            itr = 0
                            time.sleep(1)
                            while True:
                                try:
                                    arduino = serial.Serial("COM4", 9600, timeout=1)
                                    break
                                except serial.SerialException:
                                    print("Error en la comunicación, reintentando conexión")
                                    time.sleep(1)
                                    continue
                            available = 1
                            l = 0
                            print(s)
                        elapsed_time = None

                    if side == 'l':
                        if itr == 0:
                            timer = TimerThread()
                            timer.start()
                        elapsed_time = timer.get_elapsed_time()
                        if elapsed_time < 10.00:
                            degl, hip_xl, hip_yl, knee_xl, knee_yl, ankle_xl, ankle_yl = left_side(results, width,
                                                                                                   height)
                            frame2_rgb = visualization(degl, hip_xl, hip_yl, knee_xl, knee_yl, ankle_xl, ankle_yl, frame)
                        else:
                            max_angle_left = degl
                            print(max_angle_left)
                            timer.stop()
                            itr = 0
                            left = 1
                    elif side == 'r':
                        if itr == 0:
                            timer = TimerThread()
                            timer.start()
                        elapsed_time = timer.get_elapsed_time()
                        if elapsed_time < 10.00:
                            degr, hip_xr, hip_yr, knee_xr, knee_yr, ankle_xr, ankle_yr = right_side(results, width,
                                                                                                    height)
                            frame2_rgb = visualization(degr, hip_xr, hip_yr, knee_xr, knee_yr, ankle_xr, ankle_yr, frame)
                        else:
                            max_angle_right = degr
                            print(max_angle_right)
                            timer.stop()
                            itr = 0
                            right = 1
                    if side != 'c':
                        if elapsed_time is None or elapsed_time < 10.00 and itr < 10:
                            itr = itr + 1
                root.update()
                if s == -1:

                    img2 = Image.fromarray(frame_rgb)
                    img2 = ImageTk.PhotoImage(image=img2)
                    label2.config(image=img2)
                    label2.img = img2

                else:

                    img2 = Image.fromarray(frame2_rgb)
                    img2 = ImageTk.PhotoImage(image=img2)
                    label2.config(image=img2)
                    label2.img = img2

            else:

                img2 = Image.fromarray(frame_rgb)
                img2 = ImageTk.PhotoImage(image=img2)
                label2.config(image=img2)
                label2.img = img2

            if cv2.waitKey(1) & 0xFF == ord('q'):
                arduino.close()
                break
            root.update()

        root.destroy()
        arduino.close()
        cap.release()
        cv2.waitKey(0)
        cv2.destroyAllWindows()


def thread2():
    kb.Listener(pulsa).run()


t1 = threading.Thread(target=thread1)
t2 = threading.Thread(target=thread2)

t1.start()
t2.start()

